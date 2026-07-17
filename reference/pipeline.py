"""Epitome end-to-end pipeline.

New completed SurveyMonkey responses -> score -> radar -> personalised
PDF -> (approval or auto-send) -> email -> audit log.

Modes (config "mode"):
  "approve" : report is emailed to the OWNER only, who forwards it to the
              client after checking. Nothing goes to the client directly.
  "auto"    : report is emailed straight to the client; the owner gets a
              notification copy.

State: last processed response timestamp is kept in state.json so each
run only handles new responses. The audit log (Epitome_Results_Log.xlsx)
records every scored response: name, email, date, all four totals,
leading archetype(s), mode, and delivery status.

Run:  python pipeline.py config.json
"""
import json
import os
import smtplib
import sys
from datetime import datetime, timezone
from email.message import EmailMessage

from openpyxl import Workbook, load_workbook

from epitome_scoring import score_from_statements, ARCHETYPES
from epitome_radar import radar_chart
from fill_report import personalise

LOG_HEADERS = ["Date scored", "Response ID", "First name", "Last name",
               "Email", "Sovereign", "Empress", "Consort", "Seductress",
               "Leading archetype(s)", "Mode", "Delivery"]


def append_log(path, row):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(LOG_HEADERS)
    ws.append(row)
    wb.save(path)


def send_email(cfg, to, subject, body, attachment=None):
    msg = EmailMessage()
    msg["From"] = cfg["smtp"]["from"]
    msg["To"] = to
    msg["Subject"] = subject
    msg.set_content(body)
    if attachment:
        with open(attachment, "rb") as f:
            msg.add_attachment(f.read(), maintype="application",
                               subtype="pdf",
                               filename=os.path.basename(attachment))
    with smtplib.SMTP(cfg["smtp"]["host"], cfg["smtp"]["port"]) as s:
        s.starttls()
        s.login(cfg["smtp"]["username"], cfg["smtp"]["password"])
        s.send_message(msg)


def process(cfg, resp, outdir):
    """Score one response and produce its personalised report PDF."""
    result = score_from_statements(resp["statements"])
    name = f"{(resp['first_name'] or '').strip()} {(resp['last_name'] or '').strip()}".strip() \
        or f"Respondent {resp['response_id']}"
    safe = "".join(c for c in name if c.isalnum() or c in " -_").replace(" ", "_")
    chart = os.path.join(outdir, f"radar_{safe}.png")
    report = os.path.join(outdir, f"Epitome_Report_{safe}.pdf")
    radar_chart(result["inverted"], chart)
    personalise(name, result["leading_label"], chart, report,
                template=cfg["template_pdf"])
    return name, result, report


def run(config_path):
    cfg = json.load(open(config_path))
    outdir = cfg.get("output_dir", "reports")
    os.makedirs(outdir, exist_ok=True)
    state_path = cfg.get("state_file", "state.json")
    state = json.load(open(state_path)) if os.path.exists(state_path) else {}

    from surveymonkey_client import SurveyMonkeyClient
    client = SurveyMonkeyClient(cfg["surveymonkey"]["access_token"],
                                cfg["surveymonkey"]["survey_id"])

    processed = 0
    for resp in client.completed_responses(
            start_created_at=state.get("last_created_at")):
        try:
            name, result, report = process(cfg, resp, outdir)
        except ValueError as e:
            # incomplete/unscoreable: notify owner, never guess
            send_email(cfg, cfg["owner_email"],
                       "Epitome: response could not be scored",
                       f"Response {resp['response_id']} "
                       f"({resp.get('email')}) failed scoring:\n\n{e}")
            continue

        if cfg["mode"] == "approve":
            send_email(cfg, cfg["owner_email"],
                       f"APPROVAL NEEDED - Epitome report for {name}",
                       f"{name} <{resp.get('email')}> completed the "
                       f"assessment.\nLeading archetype(s): "
                       f"{result['leading_label']}\nTotals: "
                       f"{result['totals']}\n\nReport attached. Forward "
                       f"it to them once you are happy with it.",
                       attachment=report)
            delivery = "sent to owner for approval"
        else:  # auto
            send_email(cfg, resp["email"],
                       cfg.get("client_subject",
                               "Your Epitome Archetype Framework Results"),
                       cfg.get("client_body",
                               f"Dear {name},\n\nPlease find your "
                               "personalised Epitome Archetype Framework "
                               "report attached.\n\nWarm regards,\nMerle"),
                       attachment=report)
            send_email(cfg, cfg["owner_email"],
                       f"Epitome report sent to {name}",
                       f"Leading archetype(s): {result['leading_label']}\n"
                       f"Totals: {result['totals']}",
                       attachment=report)
            delivery = f"emailed to {resp['email']}"

        append_log(cfg.get("log_file", "Epitome_Results_Log.xlsx"), [
            datetime.now(timezone.utc).isoformat(timespec="seconds"),
            resp["response_id"], resp.get("first_name"),
            resp.get("last_name"), resp.get("email"),
            *[result["totals"][a] for a in ARCHETYPES],
            result["leading_label"], cfg["mode"], delivery])

        state["last_created_at"] = resp["created_at"]
        json.dump(state, open(state_path, "w"))
        processed += 1

    print(f"Processed {processed} new response(s).")


if __name__ == "__main__":
    run(sys.argv[1] if len(sys.argv) > 1 else "config.json")
